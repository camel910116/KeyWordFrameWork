#encoding=utf-8
from openpyxl import *
from openpyxl.styles import Border, Side, Font
import time
import os
from FormatTime import date_time_chinese
#所有操作excel的方法

class ParseExcel(object):
    def __init__(self,excel_file_path):
        self.excel_file_path=excel_file_path
        self.workbook=load_workbook(excel_file_path)
        self.font=Font(color=None)
        self.colorDict={"red":'FFFF3030',"green":'FF008B00'}
        self.sheet=self.get_sheet_by_index(0)

    #设置当前要操作的sheet对象，使用index来获取相应的sheet
    def set_sheet_by_index(self,sheet_index):
        self.sheet = self.get_sheet_by_index(sheet_index)
    # 设置当前要操作的sheet对象，使用sheet名称来获取相应的sheet
    def set_sheet_by_name(self,sheet_name):
        self.sheet=self.workbook.get_sheet_by_name(sheet_name)

    #获取当前默认sheet的名字
    def get_default_name(self):
        return self.sheet.title

    #通过sheet名称获取sheet对象
    def get_sheet_by_name(self,sheet_name):
        self.sheet=self.workbook.get_sheet_by_name(sheet_name)
        return self.workbook.get_sheet_by_name(sheet_name)

    # 通过sheet 索引获取sheet对象
    def get_sheet_by_index(self,sheet_index):
        sheet_name=self.workbook.get_sheet_names()[sheet_index]
        self.sheet=self.get_sheet_by_name(sheet_name)
        return self.get_sheet_by_name(sheet_name)

    #获取默认sheet中最大的行数
    def get_max_row_no(self):
        return self.sheet.max_row

    #获取默认 sheet 的最大列数
    def get_max_col_no(self):
        return self.sheet.max_column

    #获取默认sheet的最小（起始）行号
    def get_min_row_no(self):
        return self.sheet.min_row

    # 获取默认sheet的最小（起始）列号
    def get_min_col_no(self):
        return self.sheet.min_column

    # 获取默认 sheet 的所有行对象，注意：openpyxl的版本必须是2.3.5
    def get_all_rows(self):
        return self.sheet.rows

    # 获取默认sheet中的所有列对象
    def get_all_cols(self):
        return self.sheet.columns

    #从默认sheet中获取某一行，第一行从0开始
    def get_single_row(self,row_no):
        return self.get_all_rows()[row_no]

    # #从默认sheet中获取某一列，第一列从0开始
    def get_single_col(self,col_no):
        return self.get_all_cols()[col_no]

    #从默认sheet中，通过行号和列号获取指定的单元格，注意行号和列号从1开始
    def get_cell(self,row_no,col_no):
        return self.sheet.cell(row=row_no,column=col_no)

    # 从默认sheet中，通过行号和列号获取指定的单元格中的内容，注意行号和列号从1开始
    def get_cell_content(self,row_no,col_no):
        return self.sheet.cell(row=row_no, column=col_no).value

    # 从默认sheet中，通过行号和列号向指定单元格中写入指定内容，注意行号和列号从1开始
    # 调用此方法的时候，excel不要处于打开状态
    def write_cell_content(self,row_no,col_no,content,font=None):
        self.sheet.cell(row=row_no, column=col_no).value=content
        self.workbook.save(self.excel_file_path)

    # 从默认sheet中，通过行号和列号向指定单元格中写入当前日期，注意行号和列号从1开始
    #调用此方法的时候，excel不要处于打开状态,读的时候可以开着文件，写入的时候不行.
    def write_cell_current_time(self,row_no,col_no):
        self.sheet.cell(row=row_no, column=col_no).value =date_time_chinese()
        self.workbook.save(self.excel_file_path)

    def save_excel_file(self):
        self.workbook.save(self.excel_file_path)

if __name__=='__main__':
    pe=ParseExcel("e:\\test.xlsx")
    pe.set_sheet_by_index(0)
    print pe.get_default_name()
    pe.set_sheet_by_name("Sheet2")
    print pe.get_default_name()
    print pe.get_sheet_by_name("Sheet2")
    print pe.get_sheet_by_index(0)
    print pe.get_max_row_no()
    print pe.get_max_col_no()
    print pe.get_min_row_no()
    print pe.get_min_col_no()
    print pe.get_all_rows()[0]
    print len(pe.get_all_rows())
    for cell in pe.get_all_rows()[0]:
        print cell.value
    print pe.get_all_cols()
    for col in pe.get_all_cols()[0]:
        print col.value
    for cell in pe.get_single_row(0):
        print cell.value
    for col in pe.get_single_col(0):
        print col.value
    print pe.get_cell(1, 1)

    print pe.get_cell_content(1, 1)
    pe.write_cell_content(8, 8, "光荣之路")
    print pe.get_cell_content(8, 8)
    pe.write_cell_current_time(8, 8)
    print pe.get_cell_content(8, 8)
